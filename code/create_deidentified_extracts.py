#!/usr/bin/env python3
"""Create privacy-reduced CSV extracts for the walking-speed paper.

The raw workbooks are retained in data/restricted_source_workbooks for private
review. This script creates public candidate analysis extracts that drop exact
video timestamps, free-text descriptions/notes, and body-measurement estimate
columns not needed for the paper-level speed/category checks.
"""
from __future__ import annotations

import csv
import json
import math
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "restricted_source_workbooks"
OUT_DIR = ROOT / "data" / "public_deidentified"
OUT_DIR.mkdir(parents=True, exist_ok=True)

CITY_SOURCE = RAW_DIR / "city_road_bridge_source_workbook.xlsx"
BAY_SOURCE = RAW_DIR / "bay_street_source_workbook.xlsx"

def clean_header(value):
    return str(value).strip() if value is not None else ""

def as_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()

def as_float(value):
    try:
        out = float(value)
    except Exception:
        return None
    return out if math.isfinite(out) else None

def load_rows(path: Path, sheet: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = [clean_header(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in values):
            continue
        row = dict(zip(headers, values))
        row["_excel_row"] = excel_row
        rows.append(row)
    return rows

def write_csv(path: Path, rows, fields):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})

city_fields = [
    "record_id", "site", "observation_date", "collection_window", "source_sheet",
    "seconds_elapsed", "speed_mps", "walking_direction", "distance_m",
    "in_group", "using_phone", "texting", "calling", "holding_phone",
    "gender_recorded", "overtaken", "overtakes", "phone_follower",
    "phone_follower_2", "phone_follower_3", "text_follower", "group_follower"
]
city_rows = []
for raw in load_rows(CITY_SOURCE, "Raw data"):
    speed = as_float(raw.get("speed (m/s)"))
    seconds = as_float(raw.get("Seconds Elapsed"))
    if speed is None or seconds is None:
        continue
    city_rows.append({
        "record_id": f"city_road_{len(city_rows)+1:03d}",
        "site": "City Road Bridge, University of Sydney",
        "observation_date": "2018-04-24",
        "collection_window": "11:24-12:34",
        "source_sheet": "Raw data",
        "seconds_elapsed": repr(seconds),
        "speed_mps": repr(speed),
        "walking_direction": as_text(raw.get("Walking direction")),
        "distance_m": as_text(raw.get("Distance")),
        "in_group": as_text(raw.get("in group?")),
        "using_phone": as_text(raw.get("using phone?")),
        "texting": as_text(raw.get("texting")),
        "calling": as_text(raw.get("calling")),
        "holding_phone": as_text(raw.get("holding phone")),
        "gender_recorded": as_text(raw.get("Gender")),
        "overtaken": as_text(raw.get("Overtaken")),
        "overtakes": as_text(raw.get("Overtakes")),
        "phone_follower": as_text(raw.get("Phone Follower")),
        "phone_follower_2": as_text(raw.get("Phone Follower Follower")),
        "phone_follower_3": as_text(raw.get("Phone Follower Follower Follower")),
        "text_follower": as_text(raw.get("Text Follower")),
        "group_follower": as_text(raw.get("Group Follower")),
    })

bay_fields = [
    "record_id", "site", "observation_dates", "collection_windows", "source_sheet",
    "time_taken_to_cross_s", "steps_taken_for_cross", "time_for_six_steps_s",
    "people_entered_zone_count", "people_exited_zone_count", "people_in_zone_count",
    "stopped", "direction_code", "with_phone", "earphones", "following_phone_user_or_headway",
    "speed_difference_from_followed_phone_user", "gender_recorded_code", "walking_alone",
    "backpack", "carrying_one_hand", "carrying_both_hands", "heavy_weight",
    "paper_category_code", "speed_mps", "speed_uniform_step_length_mps"
]
bay_rows = []
for raw in load_rows(BAY_SOURCE, "Useful data"):
    speed = as_float(raw.get("speed measured by distance （m/s）"))
    cat = as_float(raw.get("Category (1= Uninfluenced, 2= phone user follower, 3 = phone user)"))
    if speed is None or cat not in {1.0, 2.0, 3.0}:
        continue
    bay_rows.append({
        "record_id": f"bay_street_{len(bay_rows)+1:03d}",
        "site": "Bay Street, Ultimo, NSW",
        "observation_dates": "2018-05-01; 2018-06-13",
        "collection_windows": "16:42-16:55; 11:11-11:28; 16:22-16:38",
        "source_sheet": "Useful data",
        "time_taken_to_cross_s": as_text(raw.get("Time taken to cross (s)")),
        "steps_taken_for_cross": as_text(raw.get("steps taken for cross")),
        "time_for_six_steps_s": as_text(raw.get("timetaken for 6 steps (s)")),
        "people_entered_zone_count": as_text(raw.get("Number of people had  entered the zone")),
        "people_exited_zone_count": as_text(raw.get("Number of people had exitted the zone(by enter time)")),
        "people_in_zone_count": as_text(raw.get("Numbers of people in the zone")),
        "stopped": as_text(raw.get("Stopped?")),
        "direction_code": as_text(raw.get("Direction (0= walking towards left,1=right)")),
        "with_phone": as_text(raw.get("With phone?")),
        "earphones": as_text(raw.get("Earphones?")),
        "following_phone_user_or_headway": as_text(raw.get("following phone user? Or display headway")),
        "speed_difference_from_followed_phone_user": as_text(raw.get("Speed difference from the following phone user")),
        "gender_recorded_code": as_text(raw.get("Gender 0=W 1=M")),
        "walking_alone": as_text(raw.get("walking alone?")),
        "backpack": as_text(raw.get("Backpack?")),
        "carrying_one_hand": as_text(raw.get("Carrying in one hand?")),
        "carrying_both_hands": as_text(raw.get("Carrying in both hands?")),
        "heavy_weight": as_text(raw.get("Heavy weight?")),
        "paper_category_code": str(int(cat)),
        "speed_mps": repr(speed),
        "speed_uniform_step_length_mps": as_text(raw.get("speed measured by uniform step length")),
    })

write_csv(OUT_DIR / "city_road_bridge_observations_deidentified.csv", city_rows, city_fields)
write_csv(OUT_DIR / "bay_street_observations_deidentified.csv", bay_rows, bay_fields)
summary = {
    "city_road_rows_written": len(city_rows),
    "bay_street_rows_written": len(bay_rows),
    "city_filter": "Rows with numeric seconds elapsed and speed from Raw data sheet.",
    "bay_filter": "Rows with numeric speed and paper category code 1, 2, or 3 from Useful data sheet.",
    "privacy_reductions": [
        "Dropped exact entry/exit timestamps from City Road workbook.",
        "Dropped free-text Description and Notes fields.",
        "Dropped Bay Street video-relative enter/exit times.",
        "Dropped estimated age, height, high/short, and footwear columns from public candidate extract."
    ]
}
(OUT_DIR / "extract_summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
print(json.dumps(summary, indent=2))
